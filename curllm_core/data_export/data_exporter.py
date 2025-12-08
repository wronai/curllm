import json
import csv
import io
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
from datetime import datetime


class DataExporter:
    """
    Multi-format data exporter
    
    Usage:
        exporter = DataExporter(products)
        exporter.to_json("output.json")
        exporter.to_csv("output.csv")
        exporter.to_excel("output.xlsx")
    """
    
    def __init__(self, data: Union[List[Dict], Dict], metadata: Optional[Dict] = None):
        """
        Initialize exporter
        
        Args:
            data: Data to export (list of dicts or single dict)
            metadata: Optional metadata to include
        """
        self.data = data if isinstance(data, list) else [data]
        self.metadata = metadata or {}
        self.metadata.setdefault("exported_at", datetime.now().isoformat())
        self.metadata.setdefault("count", len(self.data))
    
    def to_json(
        self, 
        file_path: Optional[Union[str, Path]] = None,
        pretty: bool = True,
        include_metadata: bool = True
    ) -> str:
        """
        Export to JSON
        
        Args:
            file_path: Optional path to save file
            pretty: Pretty print with indentation
            include_metadata: Include metadata in output
            
        Returns:
            JSON string
        """
        output = {
            "data": self.data,
        }
        
        if include_metadata:
            output["metadata"] = self.metadata
        
        json_str = json.dumps(
            output,
            indent=2 if pretty else None,
            ensure_ascii=False
        )
        
        if file_path:
            Path(file_path).write_text(json_str, encoding='utf-8')
        
        return json_str
    
    def to_jsonl(self, file_path: Optional[Union[str, Path]] = None) -> str:
        """
        Export to JSONL (JSON Lines)
        
        Args:
            file_path: Optional path to save file
            
        Returns:
            JSONL string
        """
        lines = [json.dumps(item, ensure_ascii=False) for item in self.data]
        jsonl_str = '\n'.join(lines)
        
        if file_path:
            Path(file_path).write_text(jsonl_str, encoding='utf-8')
        
        return jsonl_str
    
    def to_csv(
        self,
        file_path: Optional[Union[str, Path]] = None,
        delimiter: str = ',',
        include_headers: bool = True,
        columns: Optional[List[str]] = None
    ) -> str:
        """
        Export to CSV
        
        Args:
            file_path: Optional path to save file
            delimiter: Field delimiter
            include_headers: Include header row
            columns: Optional list of columns to include (in order)
            
        Returns:
            CSV string
        """
        if not self.data:
            return ""
        
        # Determine columns
        if columns is None:
            columns = list(self.data[0].keys()) if self.data else []
        
        # Build CSV
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=columns,
            delimiter=delimiter,
            extrasaction='ignore'
        )
        
        if include_headers:
            writer.writeheader()
        
        writer.writerows(self.data)
        
        csv_str = output.getvalue()
        
        if file_path:
            Path(file_path).write_text(csv_str, encoding='utf-8')
        
        return csv_str
    
    def to_xml(
        self,
        file_path: Optional[Union[str, Path]] = None,
        root_tag: str = "data",
        item_tag: str = "item"
    ) -> str:
        """
        Export to XML
        
        Args:
            file_path: Optional path to save file
            root_tag: Root element tag
            item_tag: Item element tag
            
        Returns:
            XML string
        """
        lines = [f'<?xml version="1.0" encoding="UTF-8"?>']
        lines.append(f'<{root_tag}>')
        
        for item in self.data:
            lines.append(f'  <{item_tag}>')
            for key, value in item.items():
                # Escape XML special characters
                value_str = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                lines.append(f'    <{key}>{value_str}</{key}>')
            lines.append(f'  </{item_tag}>')
        
        lines.append(f'</{root_tag}>')
        
        xml_str = '\n'.join(lines)
        
        if file_path:
            Path(file_path).write_text(xml_str, encoding='utf-8')
        
        return xml_str
    
    def to_markdown(
        self,
        file_path: Optional[Union[str, Path]] = None,
        columns: Optional[List[str]] = None
    ) -> str:
        """
        Export to Markdown table
        
        Args:
            file_path: Optional path to save file
            columns: Optional list of columns to include
            
        Returns:
            Markdown string
        """
        if not self.data:
            return ""
        
        # Determine columns
        if columns is None:
            columns = list(self.data[0].keys()) if self.data else []
        
        # Build table
        lines = []
        
        # Header
        header = "| " + " | ".join(columns) + " |"
        separator = "| " + " | ".join(["---"] * len(columns)) + " |"
        lines.append(header)
        lines.append(separator)
        
        # Rows
        for item in self.data:
            row = "| " + " | ".join(str(item.get(col, "")) for col in columns) + " |"
            lines.append(row)
        
        md_str = '\n'.join(lines)
        
        if file_path:
            Path(file_path).write_text(md_str, encoding='utf-8')
        
        return md_str
    
    def to_html(
        self,
        file_path: Optional[Union[str, Path]] = None,
        table_class: str = "data-table",
        include_style: bool = True
    ) -> str:
        """
        Export to HTML table
        
        Args:
            file_path: Optional path to save file
            table_class: CSS class for table
            include_style: Include basic CSS styling
            
        Returns:
            HTML string
        """
        if not self.data:
            return "<table></table>"
        
        columns = list(self.data[0].keys()) if self.data else []
        
        lines = []
        
        # Add CSS if requested
        if include_style:
            lines.append("""
<style>
.data-table {
    border-collapse: collapse;
    width: 100%;
    margin: 20px 0;
}
.data-table th, .data-table td {
    border: 1px solid #ddd;
    padding: 8px;
    text-align: left;
}
.data-table th {
    background-color: #4CAF50;
    color: white;
}
.data-table tr:nth-child(even) {
    background-color: #f2f2f2;
}
</style>
""")
        
        # Build table
        lines.append(f'<table class="{table_class}">')
        
        # Header
        lines.append('  <thead>')
        lines.append('    <tr>')
        for col in columns:
            lines.append(f'      <th>{col}</th>')
        lines.append('    </tr>')
        lines.append('  </thead>')
        
        # Body
        lines.append('  <tbody>')
        for item in self.data:
            lines.append('    <tr>')
            for col in columns:
                value = str(item.get(col, ""))
                lines.append(f'      <td>{value}</td>')
            lines.append('    </tr>')
        lines.append('  </tbody>')
        
        lines.append('</table>')
        
        html_str = '\n'.join(lines)
        
        if file_path:
            Path(file_path).write_text(html_str, encoding='utf-8')
        
        return html_str
    
    def to_yaml(self, file_path: Optional[Union[str, Path]] = None) -> str:
        """
        Export to YAML
        
        Args:
            file_path: Optional path to save file
            
        Returns:
            YAML string
        """
        try:
            import yaml
            yaml_str = yaml.dump(
                {"data": self.data, "metadata": self.metadata},
                allow_unicode=True,
                default_flow_style=False
            )
        except ImportError:
            # Fallback to simple YAML-like format
            lines = ["data:"]
            for i, item in enumerate(self.data):
                lines.append(f"  - item_{i}:")
                for key, value in item.items():
                    lines.append(f"      {key}: {value}")
            yaml_str = '\n'.join(lines)
        
        if file_path:
            Path(file_path).write_text(yaml_str, encoding='utf-8')
        
        return yaml_str
    
    def to_excel(self, file_path: Union[str, Path], sheet_name: str = "Data"):
        """
        Export to Excel (XLSX)
        
        Requires: openpyxl or xlsxwriter
        
        Args:
            file_path: Path to save file
            sheet_name: Name of the sheet
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not self.data:
                wb.save(file_path)
                return
            
            # Headers
            columns = list(self.data[0].keys())
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # Data
            for row_idx, item in enumerate(self.data, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(col_name, ""))
            
            wb.save(file_path)
            
        except ImportError:
            raise ImportError("Excel export requires 'openpyxl'. Install with: pip install openpyxl")
    
    def to_sqlite(self, db_path: Union[str, Path], table_name: str = "data"):
        """
        Export to SQLite database
        
        Args:
            db_path: Path to SQLite database file
            table_name: Name of the table
        """
        import sqlite3
        
        if not self.data:
            return
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Create table
        columns = list(self.data[0].keys())
        columns_sql = ", ".join([f"{col} TEXT" for col in columns])
        cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} ({columns_sql})")
        
        # Insert data
        placeholders = ", ".join(["?" for _ in columns])
        insert_sql = f"INSERT INTO {table_name} VALUES ({placeholders})"
        
        for item in self.data:
            values = [item.get(col, "") for col in columns]
            cursor.execute(insert_sql, values)
        
        conn.commit()
        conn.close()
